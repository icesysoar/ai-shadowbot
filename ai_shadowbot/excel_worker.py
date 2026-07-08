"""Excel/CSV 操作 Worker（F016）—— 读写 Excel/CSV 文件。

设计要点：
  - 所有方法返回统一 dict: {success: bool, data: ..., error: "..."}
  - openpyxl lazy import：仅在 read_excel / write_excel 时加载
  - csv 使用标准库 csv 模块
  - read_csv 支持自定义分隔符和 header 检测
  - 所有操作带异常保护，不抛裸异常给调用方
"""
from __future__ import annotations

import csv
import os
from typing import Any, Dict, List, Optional


class ExcelWorker:
    """Excel/CSV 文件操作 Worker。

    支持 CSV（标准库 csv）和 Excel（openpyxl，lazy import）。
    """

    @staticmethod
    def _success(data: Any = None) -> dict:
        return {"success": True, "data": data, "error": None}

    @staticmethod
    def _error(msg: str) -> dict:
        return {"success": False, "data": None, "error": msg}

    # ------------------------------------------------------------------
    # CSV 读取
    # ------------------------------------------------------------------

    def read_csv(
        self,
        path: str,
        delimiter: str = ",",
        has_header: bool = True,
    ) -> dict:
        """读取 CSV 文件。

        Args:
            path: CSV 文件路径
            delimiter: 分隔符，默认逗号
            has_header: 是否包含表头行

        Returns:
            {success, data: {headers: [...], rows: [[...],...], row_count: N}, error}
        """
        if not os.path.isfile(path):
            return self._error(f"文件不存在: {path}")

        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f, delimiter=delimiter)
                all_rows = list(reader)
        except UnicodeDecodeError:
            # 尝试 GBK 编码
            try:
                with open(path, "r", encoding="gbk", newline="") as f:
                    reader = csv.reader(f, delimiter=delimiter)
                    all_rows = list(reader)
            except Exception as e:
                return self._error(f"CSV 编码错误: {e}")
        except Exception as e:
            return self._error(f"读取 CSV 失败: {e}")

        if not all_rows:
            return self._success({
                "headers": [],
                "rows": [],
                "row_count": 0,
            })

        if has_header:
            headers = all_rows[0]
            rows = all_rows[1:]
        else:
            headers = [f"col_{i}" for i in range(len(all_rows[0]))]
            rows = all_rows

        return self._success({
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
        })

    # ------------------------------------------------------------------
    # Excel 读取（openpyxl lazy import）
    # ------------------------------------------------------------------

    def read_excel(self, path: str, sheet: str = "") -> dict:
        """读取 Excel 文件（.xlsx/.xls）。

        Args:
            path: Excel 文件路径
            sheet: 工作表名，为空则取活动工作表

        Returns:
            {success, data: {sheet_names, active_sheet, headers, rows, row_count}, error}
        """
        if not os.path.isfile(path):
            return self._error(f"文件不存在: {path}")

        try:
            import openpyxl
        except ImportError:
            return self._error("openpyxl 未安装，请 pip install openpyxl")

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            return self._error(f"打开 Excel 文件失败: {e}")

        try:
            sheet_names = wb.sheetnames
            if sheet:
                if sheet not in sheet_names:
                    wb.close()
                    return self._error(f"工作表 '{sheet}' 不存在，可用: {sheet_names}")
                ws = wb[sheet]
            else:
                ws = wb.active
                sheet = ws.title if ws else ""

            if ws is None:
                wb.close()
                return self._error("无法获取工作表")

            # 读取所有行
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row))

            if not all_rows:
                wb.close()
                return self._success({
                    "sheet_names": sheet_names,
                    "active_sheet": sheet,
                    "headers": [],
                    "rows": [],
                    "row_count": 0,
                })

            headers = all_rows[0]
            rows = all_rows[1:]

            wb.close()
            return self._success({
                "sheet_names": sheet_names,
                "active_sheet": sheet,
                "headers": headers,
                "rows": rows,
                "row_count": len(rows),
            })
        except Exception as e:
            try:
                wb.close()
            except Exception:
                pass
            return self._error(f"读取 Excel 数据失败: {e}")

    # ------------------------------------------------------------------
    # CSV 写入
    # ------------------------------------------------------------------

    def write_csv(self, path: str, headers: list, rows: list) -> dict:
        """写入 CSV 文件。

        Args:
            path: 输出 CSV 文件路径
            headers: 表头列表
            rows: 数据行列表（每行为一个 list）

        Returns:
            {success, message, error}
        """
        try:
            # 确保目录存在
            dirname = os.path.dirname(path)
            if dirname:
                os.makedirs(dirname, exist_ok=True)

            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                if headers:
                    writer.writerow(list(headers))
                for row in rows:
                    writer.writerow(list(row))

            return self._success(f"已写入 {len(rows)} 行到 {path}")
        except Exception as e:
            return self._error(f"写入 CSV 失败: {e}")

    # ------------------------------------------------------------------
    # Excel 写入（openpyxl lazy import）
    # ------------------------------------------------------------------

    def write_excel(
        self,
        path: str,
        headers: list,
        rows: list,
        sheet: str = "Sheet1",
    ) -> dict:
        """写入 Excel 文件（.xlsx）。

        Args:
            path: 输出 Excel 文件路径
            headers: 表头列表
            rows: 数据行列表
            sheet: 工作表名

        Returns:
            {success, message, error}
        """
        try:
            import openpyxl
        except ImportError:
            return self._error("openpyxl 未安装，请 pip install openpyxl")

        try:
            wb = openpyxl.Workbook()
            if sheet != "Sheet1":
                # 重命名默认 sheet
                ws = wb.active
                if ws:
                    ws.title = sheet
            else:
                ws = wb.active

            if ws is None:
                wb.close()
                return self._error("无法创建工作表")

            # 写入表头
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)

            # 写入数据
            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 确保目录存在
            dirname = os.path.dirname(path)
            if dirname:
                os.makedirs(dirname, exist_ok=True)

            wb.save(path)
            wb.close()
            return self._success(f"已写入 {len(rows)} 行到 {path}")
        except Exception as e:
            try:
                wb.close()
            except Exception:
                pass
            return self._error(f"写入 Excel 失败: {e}")

    # ------------------------------------------------------------------
    # CSV 合并
    # ------------------------------------------------------------------

    def merge_csv(self, paths: list, output: str) -> dict:
        """合并多个 CSV 文件为单个 CSV。

        所有 CSV 必须有相同的表头（按第一个文件的表头为准）。

        Args:
            paths: CSV 文件路径列表
            output: 输出合并后 CSV 的路径

        Returns:
            {success, data: {total_rows: N}, error}
        """
        if not paths:
            return self._error("paths 列表为空")

        all_rows: List[list] = []
        merged_headers: Optional[list] = None

        for i, p in enumerate(paths):
            result = self.read_csv(p)
            if not result["success"]:
                return self._error(f"读取第 {i+1} 个文件失败 ({p}): {result['error']}")

            data = result["data"]
            headers = data["headers"]
            rows = data["rows"]

            if merged_headers is None:
                merged_headers = headers
            elif headers != merged_headers:
                return self._error(
                    f"文件 {p} 的表头不一致。期望: {merged_headers}, 实际: {headers}"
                )

            all_rows.extend(rows)

        if merged_headers is None:
            merged_headers = []

        result = self.write_csv(output, merged_headers, all_rows)
        if result["success"]:
            result["data"] = {"total_rows": len(all_rows)}
        return result

    # ------------------------------------------------------------------
    # 行筛选
    # ------------------------------------------------------------------

    def filter_rows(self, path: str, column: str, value: str) -> dict:
        """按列值筛选 CSV/Excel 行。

        Args:
            path: 文件路径
            column: 筛选列名
            value: 匹配值

        Returns:
            {success, data: {headers, rows, matched_count}, error}
        """
        # 根据扩展名决定读取方式
        ext = os.path.splitext(path)[1].lower()
        if ext in (".xlsx", ".xls", ".xlsm"):
            result = self.read_excel(path)
        else:
            result = self.read_csv(path)

        if not result["success"]:
            return result

        data = result["data"]
        headers = data["headers"]
        rows = data["rows"]

        if column not in headers:
            return self._error(f"列 '{column}' 不存在。可用列: {headers}")

        col_idx = headers.index(column)
        matched = [row for row in rows if col_idx < len(row) and str(row[col_idx]) == value]

        return self._success({
            "headers": headers,
            "rows": matched,
            "matched_count": len(matched),
        })
